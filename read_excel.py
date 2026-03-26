import sys
try:
    import pandas as pd
    df = pd.read_excel(sys.argv[1], nrows=5)
    print("Columns:", df.columns.tolist())
    print("First row:", df.iloc[0].tolist())
except ImportError:
    print("Pandas not installed. Trying openpyxl...")
    from openpyxl import load_workbook
    wb = load_workbook(sys.argv[1])
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(row)
        if i >= 5: break
